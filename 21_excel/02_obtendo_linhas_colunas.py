#!/usr/bin/python3

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
# Porque usar o util https://is.gd/YrDuST

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

print('Obtendo a letra da coluna a partir de um inteiro')
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))


print('\nObtendo o numero da coluna a partir da letra')
print(get_column_letter(sheet.max_column))
# Obetendo a letra da ultima coluna
print(column_index_from_string(get_column_letter(sheet.max_column)))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))

print('\nObtendo linhas e colunas das planilhas')

# Imprimindo os valores em uma tupla
# Contem tres duplas dentro
# print(tuple(sheet['A1':'C3']))

# print(tuple(sheet.columns)[1])
# for cellObj in list(sheet.columns)[1]:
#     print(cellObj.value)

# Percorrendo a area de A1 a C3
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

print('\nAcessando valores de celulas de uma linha ou coluna')
for cellObj in list(sheet.columns)[1]:  # 1 = Coluna B
    print(cellObj.value)
    # Porque precisa utilizar o metodo list https://is.gd/I3d9PR
