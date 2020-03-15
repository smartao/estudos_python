#!/usr/bin/python3

import openpyxl

print('Opening Excel Documents with OpenPyXL')
# This Workbook object represents the Excel file,
# a bit like how a File object represents an opened text file.
wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))
print('\nGetting Sheets from the Workbook')
# import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)
print(sheet.title)
# anotherSheet = wb.get_active_sheet()  # Esta com erro
# anotherSheet  # Esta com erro

print('\nGetting Cells from the Sheets')
wb = openpyxl.load_workbook('example.xlsx')  # Carregando o arquivo excel
sheet = wb.get_sheet_by_name('Sheet1')  # Carregando a planilha (aba)
print(sheet['A1'])  # Mostrando o tipo de objeto
print(sheet['A1'].value)  # Mostrando o valor da celula sera a data
c = sheet['B2']
print('\nValor de B2:' + str(c.value))  # Mostrando o valor da celular Apples
print('Linha  ' + str(c.row) + ', Coluna ' + str(c.column) + ', e: ' + str(c.value))
print('Celula ' + str(c.coordinate) + ' e: ' + str(c.value))

print('Valor da Celula: ' + str(sheet['C1'].value))

'''
Especificar uma coluna pela letra pode ser complicado para programar
especialmente porque, após a coluna Z, as colunas começar a usar 2 letras

O inteiro correspondente a primeira linha ou coluna é 1, e não 0.
'''

print('\nTrabalhando com localização e for basico')
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):  # for que ira de 1 a 8, pulando de 2 em 2
    print(i, sheet.cell(row=i, column=2).value)

print('\nOutra forma de carregar arquivos')
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

for i in range(1, 8, 2):  # for que ira de 1 a 8, pulando de 2 em 2
    print(i, sheet.cell(row=i, column=2).value)

print('\nUltima linha com valor : ' + str(sheet.max_row))
print('Ultima coluna com valor : ' + str(sheet.max_column))
print('Retorna o ID da coluna : ' + str(cell.col_idx))
# sobre https://is.gd/kKVjhb
