#!/usr/bin/python3
# update_produce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# Os tipos de produtos e seus processo atualizados
PRICE_UPDATES = {'Garlic': 99.07,
                 'Celery': 101.19,
                 'Lemon': 333.27}

# Percorre as linhas em um loop e atualizar os preços
for row_num in range(2, sheet.max_row):  # 2 = Pulando a primeira linha
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save('updated_produce_sales.xlsx')
